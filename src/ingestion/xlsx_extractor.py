"""
XLSX text extraction using openpyxl
"""
from openpyxl import load_workbook
from pathlib import Path
from typing import Optional, Dict, List
import logging

logger = logging.getLogger(__name__)


class XLSXExtractor:
    """Extract text from XLSX files"""

    @staticmethod
    def extract_text(file_path: Path) -> Optional[str]:
        """
        Extract all text from XLSX file

        Args:
            file_path: Path to XLSX file

        Returns:
            Extracted text or None if extraction fails
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
            text_content = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text_content.append(f"=== Sheet: {sheet_name} ===\n")

                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        text_content.append(" | ".join(row_text))

            wb.close()

            full_text = "\n".join(text_content)
            return full_text if full_text.strip() else None

        except Exception as e:
            logger.error(f"Error extracting text from XLSX {file_path}: {e}")
            return None

    @staticmethod
    def extract_sheets(file_path: Path) -> Dict[str, List[List]]:
        """
        Extract all sheets as structured data

        Args:
            file_path: Path to XLSX file

        Returns:
            Dictionary mapping sheet names to row data
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
            sheets = {}

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []

                for row in sheet.iter_rows(values_only=True):
                    # Convert None to empty string and everything else to string
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    # Only add non-empty rows
                    if any(cell for cell in row_data):
                        rows.append(row_data)

                if rows:
                    sheets[sheet_name] = rows

            wb.close()
            return sheets

        except Exception as e:
            logger.error(f"Error extracting sheets from XLSX {file_path}: {e}")
            return {}

    @staticmethod
    def extract_metadata(file_path: Path) -> Dict:
        """
        Extract XLSX metadata

        Args:
            file_path: Path to XLSX file

        Returns:
            Dictionary of metadata
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)

            metadata = {
                'sheet_count': len(wb.sheetnames),
                'sheet_names': wb.sheetnames,
                'created': wb.properties.created if hasattr(wb.properties, 'created') else None,
                'modified': wb.properties.modified if hasattr(wb.properties, 'modified') else None,
                'creator': wb.properties.creator if hasattr(wb.properties, 'creator') else '',
                'last_modified_by': wb.properties.lastModifiedBy if hasattr(wb.properties, 'lastModifiedBy') else '',
            }

            # Count total rows across all sheets
            total_rows = 0
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                total_rows += sheet.max_row

            metadata['total_rows'] = total_rows

            wb.close()
            return metadata

        except Exception as e:
            logger.error(f"Error extracting XLSX metadata {file_path}: {e}")
            return {}

    @staticmethod
    def extract_sheet(file_path: Path, sheet_name: str) -> List[List]:
        """
        Extract specific sheet

        Args:
            file_path: Path to XLSX file
            sheet_name: Name of sheet to extract

        Returns:
            List of rows
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)

            if sheet_name not in wb.sheetnames:
                wb.close()
                return []

            sheet = wb[sheet_name]
            rows = []

            for row in sheet.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                if any(cell for cell in row_data):
                    rows.append(row_data)

            wb.close()
            return rows

        except Exception as e:
            logger.error(f"Error extracting sheet {sheet_name} from XLSX {file_path}: {e}")
            return []

    @staticmethod
    def get_sheet_names(file_path: Path) -> List[str]:
        """
        Get list of sheet names

        Args:
            file_path: Path to XLSX file

        Returns:
            List of sheet names
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        except Exception as e:
            logger.error(f"Error getting sheet names from XLSX {file_path}: {e}")
            return []

    @staticmethod
    def extract_as_text_blocks(file_path: Path) -> List[str]:
        """
        Extract sheets as text blocks (one per sheet)

        Args:
            file_path: Path to XLSX file

        Returns:
            List of text blocks
        """
        try:
            sheets = XLSXExtractor.extract_sheets(file_path)
            blocks = []

            for sheet_name, rows in sheets.items():
                lines = [f"Sheet: {sheet_name}"]
                for row in rows:
                    line = " | ".join(row)
                    if line.strip():
                        lines.append(line)

                blocks.append("\n".join(lines))

            return blocks

        except Exception as e:
            logger.error(f"Error extracting text blocks from XLSX {file_path}: {e}")
            return []

    @staticmethod
    def search_cells(file_path: Path, search_term: str) -> List[Dict]:
        """
        Search for term in all cells

        Args:
            file_path: Path to XLSX file
            search_term: Term to search for

        Returns:
            List of matches with sheet, row, col information
        """
        try:
            wb = load_workbook(str(file_path), data_only=True, read_only=True)
            matches = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell and search_term.lower() in str(cell).lower():
                            matches.append({
                                'sheet': sheet_name,
                                'row': row_idx,
                                'col': col_idx,
                                'value': str(cell)
                            })

            wb.close()
            return matches

        except Exception as e:
            logger.error(f"Error searching XLSX {file_path}: {e}")
            return []
